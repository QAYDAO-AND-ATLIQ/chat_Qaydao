"""
Quality Guard — Phase 2 report UI + stats + Excel export.
Mounted into app.py. Served under the /quality-guard path prefix (nginx X-Forwarded-Prefix).
All endpoints read-only from the isolated quality_guard DB.
"""
import io, datetime
from fastapi import APIRouter, Query, Response, Body
from fastapi.responses import HTMLResponse

router = APIRouter()

import os as _os
_INJECT_JS = ""
try:
    with open(_os.path.join(_os.path.dirname(__file__), "static_inject.js"), "r", encoding="utf-8") as _f:
        _INJECT_JS = _f.read()
except Exception:
    _INJECT_JS = "/* qg inject unavailable */"

@router.get("/inject.js")
async def inject_js():
    return Response(content=_INJECT_JS, media_type="application/javascript",
                    headers={"Cache-Control": "no-cache"})


# pool() is injected from app.py at import time
_pool = None
def bind_pool(p):
    global _pool
    _pool = p


@router.get("/stats")
async def stats():
    p = await _pool()
    async with p.acquire() as c:
        async def scalar(q, *a): return await c.fetchval(q, *a) or 0
        today      = await scalar("SELECT count(*) FROM qg_alerts WHERE created_at::date = now()::date")
        this_week  = await scalar("SELECT count(*) FROM qg_alerts WHERE created_at >= date_trunc('week', now())")
        this_month = await scalar("SELECT count(*) FROM qg_alerts WHERE created_at >= date_trunc('month', now())")
        high       = await scalar("SELECT count(*) FROM qg_alerts WHERE severity='high'")
        delays     = await scalar("SELECT count(*) FROM qg_alerts WHERE alert_type='response_delay'")
        notes      = await scalar("SELECT count(*) FROM qg_alerts WHERE message_direction='internal_note'")
        sup_reviewed     = await scalar("SELECT count(*) FROM qg_alerts WHERE supervisor_status='reviewed'")
        sup_not_reviewed = await scalar("SELECT count(*) FROM qg_alerts WHERE supervisor_status IS DISTINCT FROM 'reviewed'")
        top_emp    = await c.fetchrow("SELECT employee_name, count(*) n FROM qg_alerts GROUP BY employee_name ORDER BY n DESC LIMIT 1")
        top_type   = await c.fetchrow("SELECT alert_type, count(*) n FROM qg_alerts GROUP BY alert_type ORDER BY n DESC LIMIT 1")
        # improvement vs previous month
        prev_month = await scalar("SELECT count(*) FROM qg_alerts WHERE created_at >= date_trunc('month', now()) - interval '1 month' AND created_at < date_trunc('month', now())")
        delta = None
        if prev_month:
            delta = round(100.0 * (this_month - prev_month) / prev_month, 1)
    return {
        "today": today, "this_week": this_week, "this_month": this_month,
        "high": high, "delays": delays, "notes": notes,
        "sup_reviewed": sup_reviewed, "sup_not_reviewed": sup_not_reviewed,
        "top_employee": dict(top_emp) if top_emp else None,
        "top_type": dict(top_type) if top_type else None,
        "improvement_pct_vs_prev_month": delta,
    }


# ---------------------------------------------------------------------------
# Export shaping - clean Arabic report (essential columns only).
# Applied identically to CSV + Excel so files stay simple, ordered, Arabic,
# with no raw/technical/random columns.
# ---------------------------------------------------------------------------

_EXPORT_SQL = """
SELECT created_at, employee_name, conversation_id, alert_type, severity,
       is_repeated, ai_reason, suggested_correction, repeated_count,
       supervisor_status, matched_rule
FROM qg_alerts
ORDER BY created_at DESC
LIMIT 5000
"""

_EXPORT_HEADERS = [
    "\u0627\u0644\u062a\u0627\u0631\u064a\u062e",                                  # التاريخ
    "\u0627\u0633\u0645 \u0627\u0644\u0645\u0648\u0638\u0641",                        # اسم الموظف
    "\u0631\u0642\u0645 \u0627\u0644\u0645\u062d\u0627\u062f\u062b\u0629",              # رقم المحادثة
    "\u0646\u0648\u0639 \u0627\u0644\u062a\u0646\u0628\u064a\u0647",                    # نوع التنبيه
    "\u0627\u0644\u062e\u0637\u0648\u0631\u0629",                                  # الخطورة
    "\u0627\u0644\u062a\u0643\u0631\u0627\u0631",                                  # التكرار
    "\u0627\u0644\u0633\u0628\u0628",                                        # السبب
    "\u0627\u0644\u0645\u0642\u062a\u0631\u062d",                                  # المقترح
    "\u0639\u062f\u062f \u0627\u0644\u062a\u0643\u0631\u0627\u0631 \u062e\u0644\u0627\u0644 \u0623\u0633\u0628\u0648\u0639",  # عدد التكرار خلال أسبوع
    "\u062d\u0627\u0644\u0629 \u0627\u0644\u0645\u0634\u0631\u0641 (\u0627\u0644\u062a\u0646\u0628\u064a\u0647 \u0644\u0644\u0645\u0648\u0638\u0641)",  # حالة المشرف
    "\u062a\u0635\u0639\u064a\u062f \u062a\u0644\u0642\u0627\u0626\u064a",              # تصعيد تلقائي
]

_ALERT_TYPE_AR = {
    "first_response_delay":     "\u062a\u0623\u062e\u0651\u0631 \u0627\u0644\u0631\u062f \u0627\u0644\u0623\u0648\u0644",
    "missing_rating_close":     "\u0625\u063a\u0644\u0627\u0642 \u0628\u062f\u0648\u0646 \u062a\u0642\u064a\u064a\u0645",
    "missing_greeting":         "\u063a\u064a\u0627\u0628 \u0627\u0644\u062a\u062d\u064a\u0629",
    "customer_abuse":           "\u0625\u0633\u0627\u0621\u0629 \u0645\u0646 \u0627\u0644\u0639\u0645\u064a\u0644",
    "abuse":                    "\u0625\u0633\u0627\u0621\u0629 \u0645\u0646 \u0627\u0644\u0639\u0645\u064a\u0644",
    "excessive_internal_notes": "\u0645\u0644\u0627\u062d\u0638\u0627\u062a \u062f\u0627\u062e\u0644\u064a\u0629 \u0645\u0641\u0631\u0637\u0629",
    "unprofessional_note":      "\u0645\u0644\u0627\u062d\u0638\u0629 \u063a\u064a\u0631 \u0627\u062d\u062a\u0631\u0627\u0641\u064a\u0629",
    "unprofessional_reply":     "\u0631\u062f \u063a\u064a\u0631 \u0627\u062d\u062a\u0631\u0627\u0641\u064a",
}
_SEVERITY_AR = {
    "high":   "\u0639\u0627\u0644\u064a\u0629",
    "medium": "\u0645\u062a\u0648\u0633\u0637\u0629",
    "low":    "\u0645\u0646\u062e\u0641\u0636\u0629",
}
_REPEAT_YES = "\u0645\u0643\u0631\u0631"
_REPEAT_NO  = "\u063a\u064a\u0631 \u0645\u0643\u0631\u0631"
_SUP_DONE   = "\u062a\u0645 \u062a\u0646\u0628\u064a\u0647 \u0627\u0644\u0645\u0648\u0638\u0641"
_SUP_NOT    = "\u0644\u0645 \u064a\u062a\u0645 \u062a\u0646\u0628\u064a\u0647 \u0627\u0644\u0645\u0648\u0638\u0641"
_ESC_YES    = "\u062a\u0645 \u0627\u0644\u062a\u0635\u0639\u064a\u062f"
_ESC_NO     = "\u0644\u0627"


def _fmt_date(dt):
    if dt is None:
        return ""
    try:
        riyadh = dt + datetime.timedelta(hours=3)  # UTC -> KSA (no DST)
        return riyadh.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _shape_row(r):
    alert_ar = _ALERT_TYPE_AR.get(r["alert_type"], r["alert_type"] or "")
    sev_ar   = _SEVERITY_AR.get(r["severity"], r["severity"] or "")
    repeat_ar = _REPEAT_YES if r["is_repeated"] else _REPEAT_NO
    sup_ar   = _SUP_DONE if r["supervisor_status"] == "reviewed" else _SUP_NOT
    escalated = ("auto_escalated" in (r["matched_rule"] or ""))
    esc_ar   = _ESC_YES if escalated else _ESC_NO
    return [
        _fmt_date(r["created_at"]),
        r["employee_name"] or "",
        r["conversation_id"],
        alert_ar,
        sev_ar,
        repeat_ar,
        (r["ai_reason"] or "").strip(),
        (r["suggested_correction"] or "").strip(),
        r["repeated_count"] if r["repeated_count"] is not None else "",
        sup_ar,
        esc_ar,
    ]


@router.get("/report.xlsx")
async def report_xlsx():
    p = await _pool()
    async with p.acquire() as c:
        rows = await c.fetch(_EXPORT_SQL)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(content="openpyxl not installed", status_code=500)

    wb = Workbook(); ws = wb.active; ws.title = "\u062a\u0642\u0631\u064a\u0631 \u0627\u0644\u062c\u0648\u062f\u0629"
    ws.sheet_view.rightToLeft = True

    ws.append(_EXPORT_HEADERS)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E7D32")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

    for r in rows:
        ws.append(_shape_row(r))

    ws.freeze_panes = "A2"
    body_align = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_align

    widths = [17, 14, 12, 18, 10, 10, 40, 40, 12, 18, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(content=bio.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=quality_guard.xlsx"})


@router.get("/report")
async def report(date_from: str = Query(default=None), date_to: str = Query(default=None),
                 employee: str = Query(default=None), alert_type: str = Query(default=None),
                 severity: str = Query(default=None), sup_status: str = Query(default=None)):
    p = await _pool()
    q = "SELECT * FROM qg_alerts WHERE 1=1"
    args, i = [], 0
    def add(cond, val):
        nonlocal i
        i += 1; args.append(val); return f" AND {cond} ${i}"
    if date_from: q += add("created_at >=", datetime.datetime.fromisoformat(date_from))
    if date_to:   q += add("created_at <=", datetime.datetime.fromisoformat(date_to))
    if employee:  q += add("employee_email =", employee)
    if alert_type:q += add("alert_type =", alert_type)
    if severity:  q += add("severity =", severity)
    if sup_status == 'reviewed':
        q += " AND supervisor_status = 'reviewed'"
    elif sup_status == 'not_reviewed':
        q += " AND supervisor_status IS DISTINCT FROM 'reviewed'"
    q += " ORDER BY created_at DESC LIMIT 1000"
    async with p.acquire() as c:
        rows = await c.fetch(q, *args)
    return {"count": len(rows), "alerts": [dict(r) for r in rows]}


@router.get("/report.csv")
async def report_csv():
    import csv as _csv
    p = await _pool()
    async with p.acquire() as c:
        rows = await c.fetch(_EXPORT_SQL)
    buf = io.StringIO()
    buf.write("\ufeff")  # UTF-8 BOM so Excel opens Arabic correctly
    w = _csv.writer(buf)
    w.writerow(_EXPORT_HEADERS)
    for r in rows:
        w.writerow(_shape_row(r))
    return Response(content=buf.getvalue(),
                    media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=quality_guard.csv"})


@router.put("/alert/{alert_id}/supervisor-status")
async def qg_update_supervisor_status(alert_id: int, payload: dict = Body(...)):
    """Update supervisor review status (single). Values: not_reviewed | reviewed."""
    new_status = (payload.get("status") or "").strip()
    if new_status not in ("not_reviewed", "reviewed"):
        return Response(content="invalid status", status_code=400)
    actor = (payload.get("actor") or "").strip() or None
    note = (("\u062d\u062f\u0651\u062b\u0647\u0627: " + actor) if actor else None)
    p = await _pool()
    async with p.acquire() as c:
        await c.execute(
            "UPDATE qg_alerts SET supervisor_status=$2, supervisor_note=$3 WHERE id=$1",
            alert_id, new_status, note)
    return {"id": alert_id, "supervisor_status": new_status}


@router.put("/alerts/bulk-supervisor-status")
async def qg_bulk_supervisor_status(payload: dict = Body(...)):
    """Bulk-update supervisor review status for multiple alerts at once."""
    new_status = (payload.get("status") or "").strip()
    if new_status not in ("not_reviewed", "reviewed"):
        return Response(content="invalid status", status_code=400)
    ids = payload.get("ids") or []
    ids = [int(x) for x in ids if str(x).strip().isdigit()]
    if not ids:
        return Response(content="no ids", status_code=400)
    actor = (payload.get("actor") or "").strip() or None
    note = (("\u062d\u062f\u0651\u062b\u0647\u0627: " + actor) if actor else None)
    p = await _pool()
    async with p.acquire() as c:
        await c.execute(
            "UPDATE qg_alerts SET supervisor_status=$2, supervisor_note=$3 WHERE id = ANY($1::int[])",
            ids, new_status, note)
    return {"updated": len(ids), "supervisor_status": new_status}


@router.get("/", response_class=HTMLResponse)
@router.get("", response_class=HTMLResponse)
async def dashboard():
    # prefix-aware: assets/api are called relative to /quality-guard via nginx X-Forwarded-Prefix
    return HTMLResponse(_PAGE, headers={
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache", "Expires": "0"})


_PAGE = r"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>تقارير جودة الموظفين — Quality Guard</title>
<style>
  :root{
    --bg:#f7f8fa; --card:#fff; --ink:#1f2733; --muted:#6b7280; --line:#e5e7eb;
    --brand:#1f6feb; --high:#dc2626; --med:#d97706; --low:#16a34a;
  }
  *{box-sizing:border-box} body{margin:0;background:var(--bg);color:var(--ink);
    font-family:-apple-system,"Segoe UI",Tahoma,Arial,sans-serif;font-size:14px}
  .wrap{max-width:1240px;margin:0 auto;padding:28px clamp(16px,4vw,40px) 48px}
  h1{font-size:19px;margin:0 0 4px} .sub{color:var(--muted);margin:0 0 16px;font-size:13px}
  .cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:14px;margin-bottom:26px}
  .weekly-reminder{display:flex;align-items:flex-start;gap:12px;background:#fff8e6;border:1px solid #f5d78a;border-inline-start:5px solid #e0a106;border-radius:12px;padding:14px 16px;margin-bottom:20px;box-shadow:0 1px 3px rgba(224,161,6,.08)}
  .weekly-reminder .wr-icon{font-size:20px;line-height:1.4;flex-shrink:0}
  .weekly-reminder .wr-text{font-size:13.5px;line-height:1.7;color:#7a5a00}
  .weekly-reminder .wr-text strong{color:#8a4b00;font-weight:700}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px}
  .card .k{color:var(--muted);font-size:12px;margin-bottom:6px}
  .card .v{font-size:24px;font-weight:700} .card .v.sm{font-size:15px}
  .card.danger{background:#fef2f2;border-color:#fca5a5}
  .card.danger .k{color:#b91c1c} .card.danger .v{color:#dc2626}
  .filters{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px;
    display:flex;flex-wrap:wrap;gap:10px;align-items:end;margin-bottom:14px}
  .filters label{display:flex;flex-direction:column;gap:5px;font-size:12px;color:var(--muted)}
  .filters input,.filters select{padding:7px 9px;border:1px solid var(--line);border-radius:8px;font-size:13px;min-width:120px}
  .btn{background:var(--brand);color:#fff;border:0;border-radius:8px;padding:8px 14px;cursor:pointer;font-size:13px}
  .btn.ghost{background:#fff;color:var(--brand);border:1px solid var(--brand)}
  .tblwrap{background:var(--card);border:1px solid var(--line);border-radius:12px;overflow-x:auto}
  table{width:100%;border-collapse:collapse;table-layout:fixed}
  th,td{padding:8px 8px;text-align:right;border-bottom:1px solid var(--line);font-size:11.5px;vertical-align:top;word-break:break-word;overflow-wrap:anywhere}
  th{font-size:11.5px}
  th{background:#fafbfc;color:var(--muted);font-weight:600;white-space:nowrap;position:sticky;top:0;z-index:1}
  tbody tr:hover{background:#fafbfc}
  tr:last-child td{border-bottom:0}
  .cell{color:var(--ink);line-height:1.55;white-space:normal}
  .pill{display:inline-block;padding:2px 10px;border-radius:999px;font-size:11px;font-weight:600;white-space:nowrap}
  .pill.high{background:#fde8e8;color:var(--high)} .pill.medium{background:#fef3e2;color:var(--med)} .pill.low{background:#e8f5ec;color:var(--low)}
  .empty{padding:40px;text-align:center;color:var(--muted)}
  .row-actions{display:flex;gap:10px;margin-bottom:18px}
  .bulkbar{display:flex;align-items:center;gap:12px;background:#eef5ff;border:1px solid #bcd4f6;border-radius:10px;padding:10px 14px;margin-bottom:12px}
  .bulkbar .bulkcount{font-size:13px;font-weight:600;color:var(--brand)}
  td input[type=checkbox],th input[type=checkbox]{width:16px;height:16px;cursor:pointer;accent-color:var(--brand)}
  .snip{color:var(--ink);display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden;line-height:1.5;cursor:pointer}
  .snip.open{-webkit-line-clamp:unset}
  .muted{color:var(--muted)}
  .pager{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-top:14px}
  .pager .info{font-size:12.5px;color:var(--muted)}
  .pager .pages{display:flex;gap:6px;flex-wrap:wrap;align-items:center}
  .pager button{min-width:34px;height:34px;padding:0 10px;border:1px solid var(--line);background:#fff;color:var(--ink);border-radius:8px;cursor:pointer;font-size:13px}
  .pager button:hover:not(:disabled){border-color:var(--brand);color:var(--brand)}
  .pager button.active{background:var(--brand);border-color:var(--brand);color:#fff;font-weight:600}
  .pager button:disabled{opacity:.45;cursor:not-allowed}
  .pager .gap{color:var(--muted);padding:0 2px}
  .tabs{display:flex;gap:8px;margin:8px 0 4px}
  .tab{background:#fff;border:1px solid var(--line);border-radius:8px 8px 0 0;padding:8px 16px;cursor:pointer;font-size:13px}
  .tab.active{background:var(--brand);color:#fff;border-color:var(--brand)}
  .subtabs{display:flex;gap:8px;margin:8px 0 22px;flex-wrap:wrap;align-items:center;padding-bottom:14px;border-bottom:1px solid var(--line)}
  .subtab{background:#fff;border:1px solid var(--line);border-radius:8px;padding:6px 12px;cursor:pointer;font-size:12.5px}
  .subtab.active{background:#eef4ff;color:var(--brand);border-color:var(--brand)}
  .gate{background:#fff;border:1px solid var(--line);border-radius:12px;padding:20px;max-width:420px;display:flex;flex-direction:column;gap:10px}
  .gate input{padding:8px;border:1px solid var(--line);border-radius:8px}
  .err{color:var(--high);font-size:12px}
  .addbox{background:#fff;border:1px solid var(--line);border-radius:12px;padding:12px;display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:12px}
  .addbox input,.addbox select{padding:7px;border:1px solid var(--line);border-radius:8px;font-size:13px}
  .rrow{display:flex;gap:8px;align-items:center;background:#fff;border:1px solid var(--line);border-radius:8px;padding:8px 10px;margin-bottom:6px;font-size:12.5px;flex-wrap:wrap}
  .rrow .ph{font-weight:600;min-width:160px}
  .rrow .x{margin-inline-start:auto;display:flex;gap:6px}
  .admin-head{display:flex;justify-content:space-between;align-items:center;margin:6px 0 18px}
  .admin-title{font-size:18px;margin:0;font-weight:700}
  .admin-sub{font-size:12px;color:var(--muted);margin:2px 0 0}
  .section{margin-top:8px;padding-top:8px}
  .help{background:#eef4ff;border:1px solid #d6e4ff;color:#33415c;border-radius:10px;padding:14px 16px;font-size:12.5px;line-height:1.8;margin-bottom:18px}
  .toolbar{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:18px}
  .toolbar .search{flex:1;min-width:180px;padding:8px 10px;border:1px solid var(--line);border-radius:8px;font-size:13px}
  .toolbar .filter{padding:8px;border:1px solid var(--line);border-radius:8px;font-size:12.5px}
  .cardlist{display:flex;flex-direction:column;gap:12px}
  .qcard{background:#fff;border:1px solid var(--line);border-radius:12px;padding:14px 16px;display:flex;gap:14px;align-items:flex-start;flex-wrap:wrap}
  .qcard.off{opacity:.55}
  .qcard .main{flex:1;min-width:200px}
  .qcard .ph{font-weight:600;font-size:13.5px;margin-bottom:3px}
  .qcard .meta{font-size:11.5px;color:var(--muted);display:flex;gap:8px;flex-wrap:wrap}
  .qcard .sub{font-size:12px;color:var(--ink);margin-top:4px}
  .qcard .acts{display:flex;gap:6px;align-items:center}
  .chip{display:inline-block;padding:1px 8px;border-radius:999px;font-size:10.5px;font-weight:600}
  .chip.high{background:#fde8e8;color:#dc2626}.chip.medium{background:#fef3e2;color:#d97706}.chip.low{background:#e8f5ec;color:#16a34a}
  .chip.scope{background:#eef2f7;color:#475569}
  .iconbtn{background:#fff;border:1px solid var(--line);border-radius:7px;padding:5px 9px;cursor:pointer;font-size:12px}
  .iconbtn:hover{background:#f5f7fa}
  .iconbtn.danger{color:#dc2626;border-color:#f3c9c9}
  .switch{position:relative;width:40px;height:22px;cursor:pointer}
  .switch input{display:none}
  .slider{position:absolute;inset:0;background:#cbd5e1;border-radius:999px;transition:.2s}
  .slider:before{content:'';position:absolute;width:16px;height:16px;background:#fff;border-radius:50%;top:3px;right:3px;transition:.2s}
  .switch input:checked + .slider{background:#16a34a}
  .switch input:checked + .slider:before{transform:translateX(-18px)}
  .formgrid{display:flex;flex-direction:column;gap:10px;max-width:560px}
  .formgrid label{font-size:12.5px;color:var(--muted);display:flex;flex-direction:column;gap:4px}
  .formgrid input,.formgrid select,.formgrid textarea{padding:8px 10px;border:1px solid var(--line);border-radius:8px;font-size:13px;font-family:inherit}
  .modal{position:fixed;inset:0;background:rgba(15,23,42,.45);z-index:60;display:flex;align-items:center;justify-content:center;padding:16px}
  .modal-card{background:#fff;border-radius:14px;width:100%;max-width:520px;max-height:90vh;overflow:auto;box-shadow:0 20px 60px rgba(0,0,0,.25)}
  .modal-head{display:flex;justify-content:space-between;align-items:center;padding:14px 16px;border-bottom:1px solid var(--line)}
  .modal-x{background:none;border:0;font-size:16px;cursor:pointer;color:var(--muted)}
  .modal-body, #modal-body{padding:16px;display:flex;flex-direction:column;gap:10px}
  #modal-body label{font-size:12.5px;color:var(--muted);display:flex;flex-direction:column;gap:4px}
  #modal-body input,#modal-body select,#modal-body textarea{padding:8px 10px;border:1px solid var(--line);border-radius:8px;font-size:13px;font-family:inherit}
  .modal-foot{display:flex;gap:8px;justify-content:flex-start;padding:12px 16px;border-top:1px solid var(--line)}
  .empty2{padding:24px;text-align:center;color:var(--muted);font-size:13px}
  .convlink{color:var(--brand);text-decoration:none;font-weight:600}
  .convlink:hover{text-decoration:underline}
  .supsel{padding:5px 8px;border:1px solid var(--line);border-radius:7px;font-size:12px;font-family:inherit;cursor:pointer;background:#fff}
  .supsel.not_reviewed{border-color:#f0c36d;background:#fff8e8;color:#9a6700}
  .supsel.reviewed{border-color:#a7d8b4;background:#eaf7ee;color:#1a7f37}
</style>
</head>
<body>
<div class="wrap">
  <h1>🛡️ تقارير جودة الموظفين — Quality Guard</h1>
  <div class="tabs">
    <button class="tab active" id="tab-reports" onclick="showTab('reports')">📊 تقارير الجودة</button>
    <button class="tab" id="tab-settings" onclick="showTab('settings')">⚙️ إعدادات الجودة</button>
  </div>
  <p class="sub">مراقبة جودة ردود خدمة العملاء والملاحظات الداخلية · QAYDAO</p>
  <div id="panel-reports">

  <div class="weekly-reminder">
    <span class="wr-icon">📌</span>
    <div class="wr-text">
      <strong>تذكير أسبوعي للمشرف:</strong>
      يجب سحب التقرير الأسبوعي لجودة الخدمة ودمجه مع التقرير الأسبوعي لقسم خدمة العملاء ومشاركته مع الإدارة قبل الاجتماع الأسبوعي.
    </div>
  </div>

  <div class="cards" id="cards"></div>

  <div class="filters">
    <label>من تاريخ<input type="date" id="from"></label>
    <label>إلى تاريخ<input type="date" id="to"></label>
    <label>الموظف (إيميل)<input type="text" id="employee" placeholder="name@qaydao.com"></label>
    <label>نوع التنبيه
      <select id="alert_type">
        <option value="">الكل</option>
        <option value="abuse">إساءة/أسلوب</option>
        <option value="unprofessional_reply">رد غير مهني</option>
        <option value="unprofessional_note">نوت غير مهني</option>
        <option value="internal_argument">جدال داخلي</option>
        <option value="policy_risk">مخاطرة سياسة</option>
        <option value="sales_risk">مخاطرة سعرية</option>
        <option value="delay_handling_risk">تعامل مع التأخير</option>
        <option value="missing_greeting">نقص ترحيب</option>
        <option value="missing_closing_check">نقص ختام</option>
        <option value="missing_rating_close">نقص تقييم</option>
        <option value="first_response_delay">تأخر الرد الأولي</option>
        <option value="official_policy_mismatch">مخالفة سياسة رسمية</option>
        <option value="customer_abuse">إساءة من العميل</option>
        <option value="excessive_internal_notes">كثرة الملاحظات الداخلية</option>
      </select></label>
    <label>الخطورة
      <select id="severity">
        <option value="">الكل</option><option value="high">عالية</option>
        <option value="medium">متوسطة</option><option value="low">منخفضة</option>
      </select></label>
    <label>حالة المشرف
      <select id="sup_status">
        <option value="">الكل</option>
        <option value="reviewed">تم تنبيه الموظف</option>
        <option value="not_reviewed">لم يتم تنبيه الموظف</option>
      </select></label>
    <button class="btn" onclick="load()">تطبيق</button>
    <button class="btn ghost" onclick="clearF()">مسح</button>
  </div>

  <div class="row-actions">
    <a class="btn ghost" id="csv" href="report.csv">⬇ CSV</a>
    <a class="btn ghost" id="xlsx" href="report.xlsx">⬇ Excel</a>
    <button class="btn ghost" onclick="window.print()">🖨 طباعة</button>
  </div>

  <div class="bulkbar" id="bulkbar" style="display:none">
    <span class="bulkcount" id="bulkCount">0 محدد</span>
    <button class="btn" onclick="bulkNotify()">✔ تنبيه المحدّد</button>
    <button class="btn ghost" onclick="clearSel()">إلغاء التحديد</button>
  </div>

  <div class="tblwrap">
  <table id="tbl">
    <colgroup>
      <col style="width:3%"><col style="width:8%"><col style="width:9%"><col style="width:6%"><col style="width:6%">
      <col style="width:8%"><col style="width:6%"><col style="width:16%">
      <col style="width:13%"><col style="width:13%"><col style="width:4%"><col style="width:8%">
    </colgroup>
    <thead><tr>
      <th><input type="checkbox" id="selAll" onclick="toggleAll(this)" title="تحديد الكل"></th>
      <th>التاريخ/الوقت</th><th>الموظف</th><th>رقم</th><th>القناة</th>
      <th>النوع</th><th>الخطورة</th><th>المقتطف</th>
      <th>السبب</th><th>المقترح</th><th>متكرر</th><th>المشرف</th>
    </tr></thead>
    <tbody id="tb"><tr><td colspan="12" class="empty">جارٍ التحميل…</td></tr></tbody>
  </table>
  </div>
  <div class="pager" id="pager" style="display:none">
    <div class="info" id="pgInfo"></div>
    <div class="pages" id="pgPages"></div>
  </div>
</div>

  </div><!-- /panel-reports -->

  <div id="panel-settings" style="display:none">
    <div id="gate" class="gate">
      <h3>إعدادات الجودة — للمشرفين</h3>
      <p class="muted" id="gatemsg">جارٍ التحقق من صلاحياتك في Chatwoot…</p>
      <div id="gatefallback" style="display:none;flex-direction:column;gap:10px">
        <p class="err">تعذّر التحقق التلقائي من هويتك. إن كنت مشرفاً، أدخل رمز وصول Chatwoot الخاص بك (من إعدادات الملف الشخصي) أو كلمة المرور الإدارية:</p>
        <input type="text" id="cwtoken" placeholder="Chatwoot Access Token (اختياري)">
        <input type="password" id="apass" placeholder="كلمة المرور الإدارية (بديل)">
        <input type="text" id="aactor" placeholder="بريدك (للسجل) — اختياري">
        <button class="btn" onclick="adminLogin()">دخول</button>
        <span id="gateerr" class="err"></span>
      </div>
    </div>

    <div id="adminbody" style="display:none">
      <div class="admin-head">
        <div>
          <h2 class="admin-title">لوحة إدارة الجودة</h2>
          <p class="admin-sub" id="admin-actor"></p>
        </div>
        <button class="btn ghost" onclick="adminLogout()">تسجيل الخروج</button>
      </div>

      <div class="subtabs">
        <button class="subtab active" data-sub="rules" onclick="showSub('rules')">🛡️ قواعد التنبيهات</button>
        <button class="subtab" data-sub="policies" onclick="showSub('policies')">📋 السياسات الرسمية</button>
        <button class="subtab" data-sub="chat" onclick="showSub('chat')">💬 معايير الشات</button>
        <button class="subtab" data-sub="sla" onclick="showSub('sla')">⏱️ زمن الرد (SLA)</button>
        <button class="subtab" data-sub="config" onclick="showSub('config')">⚙️ إعدادات عامة</button>
        <button class="subtab" data-sub="audit" onclick="showSub('audit')">📝 سجل التدقيق</button>
      </div>

      <!-- ===== RULES ===== -->
      <div id="sub-rules" class="section">
        <div class="help">قواعد التنبيهات هي العبارات التي يرصدها النظام في ردود الموظفين أو ملاحظاتهم الداخلية. عند تطابق عبارة، يصدر تنبيه داخلي خاص بنوعها ودرجة خطورتها. يمكنك إضافة عبارات جديدة، تعديل النص أو المقترح، إيقاف قاعدة مؤقتاً، أو حذفها.</div>
        <div class="toolbar">
          <input type="text" id="rules_search" class="search" placeholder="🔍 بحث في العبارات…" oninput="renderRules()">
          <select id="rules_scope" class="filter" onchange="renderRules()">
            <option value="">كل النطاقات</option><option value="external">رد خارجي</option><option value="note">نوت داخلي</option>
          </select>
          <select id="rules_sev" class="filter" onchange="renderRules()">
            <option value="">كل الخطورات</option><option value="high">عالية</option><option value="medium">متوسطة</option><option value="low">منخفضة</option>
          </select>
          <select id="rules_sort" class="filter" onchange="renderRules()">
            <option value="phrase">ترتيب: العبارة</option><option value="severity">ترتيب: الخطورة</option><option value="alert_type">ترتيب: النوع</option>
          </select>
          <button class="btn" onclick="openRuleForm()">+ قاعدة جديدة</button>
        </div>
        <div id="ruleslist" class="cardlist"></div>
      </div>

      <!-- ===== POLICIES ===== -->
      <div id="sub-policies" class="section" style="display:none">
        <div class="help">السياسات الرسمية هي مصدر الحقيقة الذي يقارن به النظام ردود الموظفين (الشحن، الإلغاء، الاسترجاع، الضمان…). عند ذكر الموظف معلومة تخالف السياسة، يصدر تنبيه «مخالفة سياسة رسمية». أدخل النص الرسمي الصحيح والأرقام المعتمدة لكل فئة.</div>
        <div class="toolbar">
          <input type="text" id="pol_search" class="search" placeholder="🔍 بحث في السياسات…" oninput="renderPolicies()">
          <button class="btn" onclick="openPolicyForm()">+ سياسة جديدة</button>
        </div>
        <div id="policieslist" class="cardlist"></div>
      </div>

      <!-- ===== CHAT STANDARDS ===== -->
      <div id="sub-chat" class="section" style="display:none">
        <div class="help">معايير الشات تتحكم بقواعد جودة المحادثة: الترحيب في بداية المحادثة، سؤال الختام قبل الإغلاق، ورسالة التقييم. هذه مبنية على قواعد من نوع missing_greeting / missing_closing_check / missing_rating_close — ويمكنك تعديل عباراتها المقبولة هنا (تظهر كقواعد بنطاق «رد خارجي»).</div>
        <div class="toolbar">
          <input type="text" id="chat_search" class="search" placeholder="🔍 بحث…" oninput="renderChat()">
          <button class="btn" onclick="openRuleForm('chat')">+ عبارة معيار</button>
        </div>
        <div id="chatlist" class="cardlist"></div>
      </div>

      <!-- ===== SLA ===== -->
      <div id="sub-sla" class="section" style="display:none">
        <div class="help">زمن الرد الأولي (SLA): إذا راسل العميل خلال أوقات الدوام ولم يرد الموظف خلال المدة المحددة، يصدر تنبيه «تأخر الرد الأولي». حدّد المدة بالدقائق وساعات الدوام.</div>
        <div id="slabox" class="formgrid"></div>
      </div>

      <!-- ===== GENERAL CONFIG ===== -->
      <div id="sub-config" class="section" style="display:none">
        <div class="help">إعدادات عامة للنظام. عدّل القيمة ثم اضغط حفظ. التغييرات تُسجّل في سجل التدقيق.</div>
        <div id="configlist" class="cardlist"></div>
      </div>

      <!-- ===== AUDIT ===== -->
      <div id="sub-audit" class="section" style="display:none">
        <div class="help">سجل التدقيق يوثّق كل تغيير في الإعدادات: من قام به، ماذا تغيّر، والقيمة قبل وبعد.</div>
        <div class="toolbar">
          <input type="text" id="audit_search" class="search" placeholder="🔍 بحث في السجل…" oninput="renderAudit()">
          <select id="audit_action" class="filter" onchange="renderAudit()"><option value="">كل الإجراءات</option></select>
        </div>
        <div id="auditlist"></div>
      </div>
    </div>

    <!-- ===== MODAL (add/edit rule or policy) ===== -->
    <div id="qg-modal" class="modal" style="display:none">
      <div class="modal-card">
        <div class="modal-head"><b id="modal-title">—</b><button class="modal-x" onclick="closeModal()">✕</button></div>
        <div id="modal-body"></div>
        <div class="modal-foot">
          <button class="btn ghost" onclick="closeModal()">إلغاء</button>
          <button class="btn" id="modal-save" onclick="modalSave()">حفظ</button>
        </div>
      </div>
    </div>
  </div>

<script>
const SEV = {high:'عالية',medium:'متوسطة',low:'منخفضة'};
const TYPE = {abuse:'إساءة/أسلوب',unprofessional_reply:'رد غير مهني',unprofessional_note:'نوت غير مهني',internal_argument:'جدال داخلي',policy_risk:'مخاطرة سياسة',sales_risk:'مخاطرة سعرية',delay_handling_risk:'تعامل مع التأخير',missing_greeting:'نقص ترحيب',missing_closing_check:'نقص ختام',missing_rating_close:'نقص تقييم',first_response_delay:'تأخر الرد الأولي',official_policy_mismatch:'مخالفة سياسة رسمية',customer_abuse:'إساءة من العميل',excessive_internal_notes:'كثرة الملاحظات الداخلية',response_delay:'تأخر بالرد'};
const DIR = {to_customer:'للعميل',internal_note:'نوت داخلي',from_customer:'من العميل'};
const CHAN = {'Channel::WebWidget':'دردشة الموقع','Channel::Api':'واتساب/API','Channel::Whatsapp':'واتساب','Channel::Email':'بريد إلكتروني','Channel::TwitterProfile':'تويتر','Channel::FacebookPage':'فيسبوك','Channel::Telegram':'تيليجرام','Channel::Sms':'رسائل SMS','Channel::Line':'لاين'};
const SUPSTAT = {not_reviewed:'لم تتم المراجعة',reviewed:'تمت المراجعة والتنبيه'};
function chanAr(c){ return CHAN[c] || (c||'—'); }
function supAr(s){ return SUPSTAT[s] || 'لم تتم المراجعة'; }
function supSelect(a){
  var cur = (a.supervisor_status==='reviewed') ? 'reviewed' : 'not_reviewed';
  var o1 = '<option value="not_reviewed"' + (cur==='not_reviewed'?' selected':'') + '>لم تتم المراجعة</option>';
  var o2 = '<option value="reviewed"' + (cur==='reviewed'?' selected':'') + '>تمت المراجعة والتنبيه</option>';
  return '<select class="supsel ' + cur + '" onchange="setSupStatus(' + a.id + ', this)">' + o1 + o2 + '</select>';
}
// base URL for Chatwoot conversation links
var CW_ORIGIN = 'https://chat.qaydao.com';
function convLink(accountId, convId){
  var acc = accountId || 1;
  return CW_ORIGIN + '/app/accounts/' + acc + '/conversations/' + convId;
}
async function setSupStatus(alertId, sel){
  var status = sel.value;
  try {
    await fetch('alert/' + alertId + '/supervisor-status', {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({status: status, actor: (typeof ACTOR!=='undefined'? ACTOR : '')})
    });
    sel.classList.remove('not_reviewed','reviewed');
    sel.classList.add(status);
  } catch(e){ alert('تعذّر تحديث الحالة'); }
}
function qs(){
  const p=new URLSearchParams();
  const f=id=>document.getElementById(id).value;
  if(f('from')) p.set('date_from',f('from'));
  if(f('to')) p.set('date_to',f('to'));
  if(f('employee')) p.set('employee',f('employee'));
  if(f('alert_type')) p.set('alert_type',f('alert_type'));
  if(f('severity')) p.set('severity',f('severity'));
  if(f('sup_status')) p.set('sup_status',f('sup_status'));
  return p.toString();
}
async function loadCards(){
  const r=await fetch('stats'); const s=await r.json();
  const c=[
    ['تنبيهات اليوم',s.today],['هذا الأسبوع',s.this_week],['هذا الشهر',s.this_month],
    ['عالية الخطورة',s.high],['تأخر بالرد',s.delays],['نوت داخلي',s.notes],
    ['تم تنبيه الموظف', (s.sup_reviewed||0), false, true],
    ['لم يتم تنبيه الموظف', (s.sup_not_reviewed||0), false, true],
    ['أكثر موظف', s.top_employee? s.top_employee.employee_name+' ('+s.top_employee.n+')':'—', true],
    ['أكثر نوع', s.top_type? (TYPE[s.top_type.alert_type]||s.top_type.alert_type)+' ('+s.top_type.n+')':'—', true],
    ['مقابل الشهر السابق', s.improvement_pct_vs_prev_month==null?'—':(s.improvement_pct_vs_prev_month>0?'+':'')+s.improvement_pct_vs_prev_month+'%', true],
  ];
  document.getElementById('cards').innerHTML=c.map(([k,v,sm,danger])=>
    `<div class="card${danger?' danger':''}"><div class="k">${k}</div><div class="v ${sm?'sm':''}">${v}</div></div>`).join('');
}
var QG_ROWS=[];        // كل التنبيهات المطابقة (محمّلة)
var QG_PAGE=1;         // الصفحة الحالية
var QG_PER=20;         // 20 تقرير/صفحة
var QG_SEL=new Set();  // معرّفات الصفوف المحدّدة

function pageIds(){
  const start=(QG_PAGE-1)*QG_PER;
  return QG_ROWS.slice(start, start+QG_PER).map(a=>a.id);
}
function onRowSel(id, el){
  if(el.checked) QG_SEL.add(id); else QG_SEL.delete(id);
  syncSelUI();
}
function toggleAll(el){
  const ids=pageIds();
  if(el.checked) ids.forEach(i=>QG_SEL.add(i));
  else ids.forEach(i=>QG_SEL.delete(i));
  document.querySelectorAll('.rowsel').forEach(cb=>{ cb.checked = QG_SEL.has(parseInt(cb.value)); });
  syncSelUI();
}
function clearSel(){
  QG_SEL.clear();
  document.querySelectorAll('.rowsel').forEach(cb=>cb.checked=false);
  const sa=document.getElementById('selAll'); if(sa){sa.checked=false;sa.indeterminate=false;}
  syncSelUI();
}
function syncSelUI(){
  const n=QG_SEL.size;
  const bar=document.getElementById('bulkbar');
  bar.style.display = n>0 ? 'flex' : 'none';
  document.getElementById('bulkCount').textContent = n + ' محدد';
  // حالة checkbox "تحديد الكل" حسب صفحة العرض الحالية
  const ids=pageIds(); const sel=ids.filter(i=>QG_SEL.has(i)).length;
  const sa=document.getElementById('selAll');
  if(sa){ sa.checked = ids.length>0 && sel===ids.length; sa.indeterminate = sel>0 && sel<ids.length; }
}
async function bulkNotify(){
  const ids=[...QG_SEL];
  if(!ids.length) return;
  if(!confirm('تأكيد: تغيير حالة '+ids.length+' تنبيه إلى «تم تنبيه الموظف»؟')) return;
  try{
    const r=await fetch('alerts/bulk-supervisor-status',{
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({ids:ids, status:'reviewed', actor:(typeof ACTOR!=='undefined'?ACTOR:'')})
    });
    if(!r.ok) throw new Error(await r.text());
    // حدّث البيانات المحمّلة محلياً
    QG_ROWS.forEach(a=>{ if(QG_SEL.has(a.id)) a.supervisor_status='reviewed'; });
    QG_SEL.clear();
    renderPage(); loadCards();
  }catch(e){ alert('تعذّر تنبيه المحدّد: '+e.message); }
}

async function load(){
  loadCards();
  const q=qs();
  document.getElementById('csv').href='report.csv';
  document.getElementById('xlsx').href='report.xlsx';
  const tb=document.getElementById('tb');
  tb.innerHTML='<tr><td colspan="12" class="empty">جارٍ التحميل…</td></tr>';
  const r=await fetch('report'+(q?'?'+q:'')); const d=await r.json();
  QG_ROWS=(d&&d.alerts)?d.alerts:[];
  QG_PAGE=1;
  renderPage();
}

function renderPage(){
  const tb=document.getElementById('tb');
  const pager=document.getElementById('pager');
  if(!QG_ROWS.length){
    tb.innerHTML='<tr><td colspan="12" class="empty">لا توجد تنبيهات مطابقة</td></tr>';
    pager.style.display='none';
    return;
  }
  const total=QG_ROWS.length;
  const pages=Math.ceil(total/QG_PER);
  if(QG_PAGE>pages) QG_PAGE=pages;
  if(QG_PAGE<1) QG_PAGE=1;
  const start=(QG_PAGE-1)*QG_PER;
  const slice=QG_ROWS.slice(start, start+QG_PER);
  tb.innerHTML=slice.map(a=>{
    const dt=new Date(a.created_at).toLocaleString('ar-SA');
    const sev=`<span class="pill ${a.severity}">${SEV[a.severity]||a.severity}</span>`;
    const rep=a.is_repeated?`نعم (${a.repeated_count})`:'لا';
    return `<tr>
      <td><input type="checkbox" class="rowsel" value="${a.id}" ${QG_SEL.has(a.id)?'checked':''} onclick="onRowSel(${a.id},this)"></td>
      <td class="muted">${dt}</td>
      <td>${a.employee_name||'—'}</td>
      <td><a href="${convLink(a.account_id, a.conversation_id)}" target="_blank" rel="noopener" class="convlink">#${a.conversation_id}</a></td>
      <td>${chanAr(a.channel_type)}</td>
      <td>${TYPE[a.alert_type]||a.alert_type}</td>
      <td>${sev}</td>
      <td><div class="cell">${(a.message_snippet||'').replace(/</g,'&lt;')}</div></td>
      <td><div class="cell muted">${(a.ai_reason||'').replace(/</g,'&lt;')}</div></td>
      <td><div class="cell muted">${(a.suggested_correction||'').replace(/</g,'&lt;')}</div></td>
      <td>${rep}</td>
      <td>${supSelect(a)}</td>
    </tr>`;
  }).join('');
  const from=start+1, to=Math.min(start+QG_PER,total);
  document.getElementById('pgInfo').textContent=`عرض ${from}–${to} من ${total}`;
  buildPager(pages);
  pager.style.display='flex';
  syncSelUI();
}

function gotoPage(p){ QG_PAGE=p; renderPage(); document.getElementById('tbl').scrollIntoView({behavior:'smooth',block:'nearest'}); }

function buildPager(pages){
  const box=document.getElementById('pgPages');
  const cur=QG_PAGE;
  let h='';
  h+=`<button onclick="gotoPage(${cur-1})" ${cur<=1?'disabled':''}>‹ السابق</button>`;
  // نافذة أرقام مختصرة: 1 … cur-1 cur cur+1 … pages
  const nums=[];
  const push=n=>{ if(n>=1&&n<=pages&&nums.indexOf(n)<0) nums.push(n); };
  push(1); push(2);
  push(cur-1); push(cur); push(cur+1);
  push(pages-1); push(pages);
  nums.sort((a,b)=>a-b);
  let prev=0;
  nums.forEach(n=>{
    if(prev && n-prev>1) h+='<span class="gap">…</span>';
    h+=`<button class="${n===cur?'active':''}" onclick="gotoPage(${n})">${n}</button>`;
    prev=n;
  });
  h+=`<button onclick="gotoPage(${cur+1})" ${cur>=pages?'disabled':''}>التالي ›</button>`;
  box.innerHTML=h;
}
function clearF(){['from','to','employee','alert_type','severity','sup_status'].forEach(i=>document.getElementById(i).value='');load();}
load();
// honor ?tab= from the injector
try {
  var _qtab = new URLSearchParams(location.search).get('tab');
  if (_qtab === 'settings') { showTab('settings'); }
} catch(_) {}

// ---------- tabs ----------
function showTab(t){
  document.getElementById('panel-reports').style.display = (t==='reports')?'block':'none';
  document.getElementById('panel-settings').style.display = (t==='settings')?'block':'none';
  document.getElementById('tab-reports').classList.toggle('active', t==='reports');
  document.getElementById('tab-settings').classList.toggle('active', t==='settings');
}
// ---------- admin auth (passphrase kept only in memory) ----------
let CW_USER = null, CW_TOKEN = null, ADMIN = null, ACTOR = 'admin';
let SESS = null;  // {access_token, client, uid} from Chatwoot session cookie
function H(){
  const h = {'Content-Type':'application/json'};
  if(SESS){ h['X-Cw-At']=SESS.access_token; h['X-Cw-Client']=SESS.client; h['X-Cw-Uid']=SESS.uid; }
  if(CW_USER) h['X-Cw-User'] = String(CW_USER);
  if(CW_TOKEN) h['X-Cw-Token'] = CW_TOKEN;
  if(ADMIN) h['X-QG-Admin'] = ADMIN;
  return h;
}
// read a cookie by name (same-origin: chat.qaydao.com)
function getCookie(name){
  const parts = ('; '+document.cookie).split('; '+name+'=');
  if(parts.length===2) return decodeURIComponent(parts.pop().split(';').shift());
  return null;
}
// extract Chatwoot session auth headers from the cw_d_session_info cookie
function readChatwootSession(){
  try {
    const raw = getCookie('cw_d_session_info');
    if(!raw) return null;
    const hh = JSON.parse(raw);
    const at = hh['access-token'] || hh['access_token'];
    const cl = hh['client'];
    const ud = hh['uid'];
    if(at && cl && ud) return {access_token:at, client:cl, uid:ud};
  } catch(_){}
  return null;
}
// Auto-verify on load using the viewer's own Chatwoot session — no input needed
async function autoGate(){
  SESS = readChatwootSession();
  if(SESS){
    const r = await fetch('admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(SESS)});
    const d = await r.json();
    if(d.ok){ ACTOR = d.actor || ACTOR; openAdmin(); return; }
    document.getElementById('gatemsg').textContent='هذه اللوحة متاحة لمشرفي Chatwoot فقط. حسابك الحالي ليس مشرفاً.';
    SESS = null; showFallback(); return;
  }
  // also accept dashboard-app context if present (when opened inside a conversation)
  // fall through to fallback if neither is available
  document.getElementById('gatemsg').textContent='تعذّر قراءة جلسة Chatwoot تلقائياً.';
  showFallback();
}
// also listen for dashboard-app context (covers the in-conversation case)
window.addEventListener('message', function(e){
  try {
    const m = typeof e.data === 'string' ? JSON.parse(e.data) : e.data;
    if(m && m.event === 'appContext' && m.data && m.data.currentAgent && !SESS){
      CW_USER = m.data.currentAgent.id;
      ACTOR = m.data.currentAgent.email || ('user:'+CW_USER);
      fetch('admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cw_user:CW_USER})})
        .then(r=>r.json()).then(d=>{ if(d.ok){ ACTOR=d.actor||ACTOR; openAdmin(); } });
    }
  } catch(_){}
});
try { window.parent.postMessage('chatwoot-dashboard-app:fetch-info','*'); } catch(_){}
// kick off auto-verification immediately
autoGate();
function showFallback(){ document.getElementById('gatefallback').style.display='flex'; }
function openAdmin(){ document.getElementById('gate').style.display='none'; document.getElementById('adminbody').style.display='block'; setActor(); showSub('rules'); }

async function adminLogin(){
  // manual fallback: chatwoot token OR passphrase
  CW_TOKEN = document.getElementById('cwtoken').value || null;
  const pass = document.getElementById('apass').value;
  ACTOR = document.getElementById('aactor').value || ACTOR;
  const body = CW_TOKEN ? {cw_token:CW_TOKEN} : {passphrase:pass};
  const r = await fetch('admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
  const d = await r.json();
  if(d.ok){ if(!CW_TOKEN){ ADMIN = pass; } ACTOR = d.actor || ACTOR; openAdmin(); }
  else { document.getElementById('gateerr').textContent='تعذّر التحقق. تأكد من الرمز أو كلمة المرور.'; }
}
function adminLogout(){ ADMIN=null; CW_TOKEN=null; document.getElementById('gate').style.display='flex'; document.getElementById('adminbody').style.display='none'; }
// ===== professional admin panel logic =====
var RULES = [], POLICIES = [], AUDIT = [];

function showSub(x){
  ['rules','policies','chat','sla','config','audit'].forEach(function(k){
    var el = document.getElementById('sub-'+k);
    if(el) el.style.display = (k===x)?'block':'none';
  });
  document.querySelectorAll('.subtab').forEach(function(b){ b.classList.toggle('active', b.getAttribute('data-sub')===x); });
  if(x==='rules'||x==='chat') loadRules();
  if(x==='policies') loadPolicies();
  if(x==='sla') loadSla();
  if(x==='config') loadConfig();
  if(x==='audit') loadAudit();
}

function setActor(){ var e=document.getElementById('admin-actor'); if(e) e.textContent = 'مسجّل الدخول: ' + (ACTOR||'مشرف'); }

// ---------- RULES ----------
var SEV_AR = {high:'عالية',medium:'متوسطة',low:'منخفضة'};
var SCOPE_AR = {external:'رد خارجي',note:'نوت داخلي'};
// خرائط ترجمة العرض (المفاتيح الداخلية تبقى كما هي — هذه للعرض فقط)
var POLCAT_AR = {
  shipping_policy:'سياسة الشحن', delivery_time_policy:'سياسة مدة التوصيل',
  cancellation_policy:'سياسة الإلغاء', return_policy:'سياسة الإرجاع',
  refund_policy:'سياسة الاسترداد', warranty_policy:'سياسة الضمان',
  installation_policy:'سياسة التركيب', payment_policy:'سياسة الدفع',
  pricing_policy:'سياسة الأسعار', offers_policy:'سياسة العروض'
};
var SETTING_AR = {
  sla_minutes:'مدة الرد الأولي (دقائق)', work_start_hour:'بداية الدوام (الساعة)',
  work_end_hour:'نهاية الدوام (الساعة)', excluded_bot_user_ids:'حسابات مستبعدة (معرّفات البوتات)',
  excluded_sender_types:'أنواع مرسلين مستبعدة', max_internal_notes:'حد الملاحظات الداخلية في المحادثة', admin_pass_hash:'بصمة كلمة مرور الإدارة'
};
var ACTION_AR = {
  create_rule:'إضافة قاعدة', update_rule:'تعديل قاعدة', delete_rule:'حذف قاعدة',
  create_policy:'إضافة سياسة', update_policy:'تعديل سياسة', toggle_policy:'تشغيل/إيقاف سياسة',
  delete_policy:'حذف سياسة', update_setting:'تعديل إعداد'
};
var ENTITY_AR = { rules:'القواعد', policies:'السياسات', settings:'الإعدادات' };
function polcatAr(c){ return POLCAT_AR[c] || c; }
function settingAr(k){ return SETTING_AR[k] || k; }
function actionAr(a){ return ACTION_AR[a] || a; }
function entityAr(e){ return ENTITY_AR[e] || e; }
function typeAr(t){ return (typeof TYPE!=='undefined' && TYPE[t]) ? TYPE[t] : t; }
var CHAT_TYPES = ['missing_greeting','missing_closing_check','missing_rating_close'];

async function loadRules(){
  var r = await fetch('admin/rules',{headers:H()});
  if(!r.ok){ return; }
  RULES = (await r.json()).rules || [];
  renderRules(); renderChat();
}
function rulesFilterSort(list, opts){
  var q=(opts.q||'').trim();
  var out = list.filter(function(x){ return x.is_active; });
  if(opts.scope) out = out.filter(function(x){ return x.scope===opts.scope; });
  if(opts.sev) out = out.filter(function(x){ return x.severity===opts.sev; });
  if(opts.chatOnly) out = out.filter(function(x){ return CHAT_TYPES.indexOf(x.alert_type)>=0; });
  if(opts.notChat) out = out.filter(function(x){ return CHAT_TYPES.indexOf(x.alert_type)<0; });
  if(q) out = out.filter(function(x){ return (x.phrase||'').indexOf(q)>=0 || (x.alert_type||'').indexOf(q)>=0; });
  var sort = opts.sort||'phrase';
  var sevOrd={high:0,medium:1,low:2};
  out.sort(function(a,b){
    if(sort==='severity') return (sevOrd[a.severity]||9)-(sevOrd[b.severity]||9);
    if(sort==='alert_type') return (a.alert_type||'').localeCompare(b.alert_type||'');
    return (a.phrase||'').localeCompare(b.phrase||'','ar');
  });
  return out;
}
function ruleCard(x){
  return '<div class="qcard">'+
    '<div class="main">'+
      '<div class="ph">'+esc(x.phrase)+'</div>'+
      '<div class="meta"><span class="chip scope">'+(SCOPE_AR[x.scope]||x.scope)+'</span>'+
        '<span class="chip '+x.severity+'">'+(SEV_AR[x.severity]||x.severity)+'</span>'+
        '<span>'+esc(typeAr(x.alert_type))+'</span></div>'+
      (x.suggested_correction?'<div class="sub">💡 '+esc(x.suggested_correction)+'</div>':'')+
    '</div>'+
    '<div class="acts">'+
      switchHtml(x.is_active, "toggleRule("+x.id+")")+
      '<button class="iconbtn" onclick=\'editRule('+x.id+')\'>تعديل</button>'+
      '<button class="iconbtn danger" onclick="delRule("+x.id+")">حذف</button>'+
    '</div></div>';
}
function switchHtml(on, onchange){
  return '<label class="switch"><input type="checkbox" '+(on?'checked':'')+' onchange="'+onchange+'"><span class="slider"></span></label>';
}
function renderRules(){
  var opts={q:val('rules_search'),scope:val('rules_scope'),sev:val('rules_sev'),sort:val('rules_sort'),notChat:true};
  var list=rulesFilterSort(RULES,opts);
  var el=document.getElementById('ruleslist');
  el.innerHTML = list.length? list.map(ruleCard).join('') : '<div class="empty2">لا توجد قواعد مطابقة</div>';
}
function renderChat(){
  var opts={q:val('chat_search'),chatOnly:true,sort:'alert_type'};
  var list=rulesFilterSort(RULES,opts);
  var el=document.getElementById('chatlist');
  if(el) el.innerHTML = list.length? list.map(ruleCard).join('') : '<div class="empty2">لا توجد عبارات معايير بعد — أضف عبارة ترحيب/ختام/تقييم</div>';
}
async function toggleRule(id){
  var x=RULES.find(function(r){return r.id===id;}); if(!x) return;
  await fetch('admin/rules/'+id,{method:'PUT',headers:H(),body:JSON.stringify({is_active:!x.is_active,_actor:ACTOR})});
  loadRules();
}
async function delRule(id){
  if(!confirm('حذف هذه القاعدة؟')) return;
  await fetch('admin/rules/'+id+'?actor='+encodeURIComponent(ACTOR),{method:'DELETE',headers:H()});
  loadRules();
}

// ---------- MODAL (rule form) ----------
var MODAL_MODE=null, MODAL_ID=null;
function openRuleForm(kind){
  MODAL_MODE = (kind==='chat')?'rule_chat':'rule'; MODAL_ID=null;
  document.getElementById('modal-title').textContent = 'قاعدة جديدة';
  document.getElementById('modal-body').innerHTML = ruleFormHtml({scope:(kind==='chat'?'external':'external'),severity:'medium'});
  openModal();
}
function editRule(id){
  var x=RULES.find(function(r){return r.id===id;}); if(!x) return;
  MODAL_MODE='rule'; MODAL_ID=id;
  document.getElementById('modal-title').textContent = 'تعديل قاعدة';
  document.getElementById('modal-body').innerHTML = ruleFormHtml(x);
  openModal();
}
function ruleFormHtml(x){
  x=x||{};
  return ''+
    '<label>العبارة<input type="text" id="m_phrase" value="'+esc(x.phrase||'')+'"></label>'+
    '<label>النطاق<select id="m_scope">'+opts(['external','note'],['رد خارجي','نوت داخلي'],x.scope)+'</select></label>'+
    '<label>نوع التنبيه (المعرّف الداخلي)<input type="text" id="m_type" value="'+esc(x.alert_type||'')+'" placeholder="مثال: abuse، unprofessional_note، missing_greeting"></label>'+
    '<label>الخطورة<select id="m_sev">'+opts(['high','medium','low'],['عالية','متوسطة','منخفضة'],x.severity)+'</select></label>'+
    '<label>السبب (يظهر في التنبيه)<textarea id="m_reason" rows="2">'+esc(x.ai_reason||'')+'</textarea></label>'+
    '<label>المقترح البديل<textarea id="m_sugg" rows="2">'+esc(x.suggested_correction||'')+'</textarea></label>';
}

// ---------- POLICIES ----------
var POLCATS=['shipping_policy','delivery_time_policy','cancellation_policy','return_policy','refund_policy','warranty_policy','installation_policy','payment_policy','pricing_policy','offers_policy'];
async function loadPolicies(){
  var r=await fetch('admin/policies',{headers:H()}); if(!r.ok) return;
  POLICIES=(await r.json()).policies||[]; renderPolicies();
}
function renderPolicies(){
  var q=(val('pol_search')||'').trim();
  var list=POLICIES.filter(function(p){return p.is_active||true;});
  if(q) list=list.filter(function(p){return (p.policy_category||'').indexOf(q)>=0||(p.official_statement||'').indexOf(q)>=0;});
  var el=document.getElementById('policieslist');
  el.innerHTML = list.length? list.map(function(p){
    return '<div class="qcard'+(p.is_active?'':' off')+'">'+
      '<div class="main"><div class="ph">'+esc(polcatAr(p.policy_category))+(p.numbers_or_limits?' · أرقام: '+esc(p.numbers_or_limits):'')+'</div>'+
      '<div class="sub">'+esc((p.official_statement||'').slice(0,140))+'</div>'+
      (p.source_url?'<div class="meta"><span>🔗 '+esc(p.source_url.slice(0,60))+'</span></div>':'')+'</div>'+
      '<div class="acts">'+switchHtml(p.is_active,"togglePolicy("+p.id+")")+
      '<button class="iconbtn" onclick=\'editPolicy('+p.id+')\'>تعديل</button>'+
      '<button class="iconbtn danger" onclick="delPolicy("+p.id+")">حذف</button></div></div>';
  }).join('') : '<div class="empty2">لا توجد سياسات بعد — أضف النص الرسمي لكل فئة</div>';
}
function openPolicyForm(){
  MODAL_MODE='policy'; MODAL_ID=null;
  document.getElementById('modal-title').textContent='سياسة رسمية جديدة';
  document.getElementById('modal-body').innerHTML=policyFormHtml({});
  openModal();
}
function editPolicy(id){
  var p=POLICIES.find(function(x){return x.id===id;}); if(!p) return;
  MODAL_MODE='policy'; MODAL_ID=id;
  document.getElementById('modal-title').textContent='تعديل سياسة';
  document.getElementById('modal-body').innerHTML=policyFormHtml(p);
  openModal();
}
function policyFormHtml(p){
  p=p||{};
  var opts2=POLCATS.map(function(c){return '<option value="'+c+'" '+(p.policy_category===c?'selected':'')+'>'+polcatAr(c)+'</option>';}).join('');
  return ''+
    '<label>الفئة<select id="m_cat">'+opts2+'</select></label>'+
    '<label>النص الرسمي الصحيح<textarea id="m_stmt" rows="3">'+esc(p.official_statement||'')+'</textarea></label>'+
    '<label>الأرقام/الحدود المعتمدة (مثل 300)<input type="text" id="m_nums" value="'+esc(p.numbers_or_limits||'')+'"></label>'+
    '<label>رابط الصفحة الرسمية<input type="text" id="m_url" value="'+esc(p.source_url||'')+'"></label>';
}
async function togglePolicy(id){ await fetch('admin/policies/'+id+'/toggle',{method:'PUT',headers:H()}); loadPolicies(); }
async function delPolicy(id){ if(!confirm('حذف هذه السياسة؟'))return; await fetch('admin/policies/'+id+'?actor='+encodeURIComponent(ACTOR),{method:'DELETE',headers:H()}); loadPolicies(); }

// ---------- SLA ----------
var SLA_KEYS=[['sla_minutes','مدة الرد الأولي (دقائق)'],['work_start_hour','بداية الدوام (ساعة 0-23)'],['work_end_hour','نهاية الدوام (ساعة 0-23)']];
async function loadSla(){
  var r=await fetch('admin/settings',{headers:H()}); if(!r.ok) return;
  var sett=(await r.json()).settings||[];
  var map={}; sett.forEach(function(s){map[s.key]=s.value;});
  var box=document.getElementById('slabox');
  box.innerHTML = SLA_KEYS.map(function(k){
    return '<label>'+k[1]+'<input type="number" id="sla_'+k[0]+'" value="'+esc(map[k[0]]||'')+'"></label>';
  }).join('') + '<div><button class="btn" onclick="saveSla()">حفظ إعدادات SLA</button></div>';
}
async function saveSla(){
  for(var i=0;i<SLA_KEYS.length;i++){
    var key=SLA_KEYS[i][0]; var v=val('sla_'+key);
    await fetch('admin/settings/'+key,{method:'PUT',headers:H(),body:JSON.stringify({value:v,_actor:ACTOR})});
  }
  alert('تم حفظ إعدادات زمن الرد'); loadSla();
}

// ---------- CONFIG ----------
async function loadConfig(){
  var r=await fetch('admin/settings',{headers:H()}); if(!r.ok) return;
  var sett=(await r.json()).settings||[];
  var el=document.getElementById('configlist');
  el.innerHTML = sett.map(function(s){
    return '<div class="qcard"><div class="main"><div class="ph">'+esc(settingAr(s.key))+'</div><div class="meta"><span class="muted" style="font-size:10.5px">'+esc(s.key)+'</span></div></div>'+
      '<div class="acts"><input type="text" id="cfg_'+s.key+'" value="'+esc(s.value||'')+'" style="padding:7px;border:1px solid var(--line);border-radius:7px">'+
      '<button class="iconbtn" onclick="saveCfg(\''+s.key+'\')">حفظ</button></div></div>';
  }).join('');
}
async function saveCfg(key){ await fetch('admin/settings/'+key,{method:'PUT',headers:H(),body:JSON.stringify({value:val('cfg_'+key),_actor:ACTOR})}); loadConfig(); }

// ---------- AUDIT ----------
async function loadAudit(){
  var r=await fetch('admin/audit',{headers:H()}); if(!r.ok) return;
  AUDIT=(await r.json()).audit||[];
  var sel=document.getElementById('audit_action');
  var actions=Array.from(new Set(AUDIT.map(function(a){return a.action;})));
  sel.innerHTML='<option value="">كل الإجراءات</option>'+actions.map(function(a){return '<option value="'+a+'">'+actionAr(a)+'</option>';}).join('');
  renderAudit();
}
function renderAudit(){
  var q=(val('audit_search')||'').trim(); var act=val('audit_action');
  var list=AUDIT.slice();
  if(act) list=list.filter(function(a){return a.action===act;});
  if(q) list=list.filter(function(a){return JSON.stringify(a).indexOf(q)>=0;});
  var el=document.getElementById('auditlist');
  el.innerHTML='<table><thead><tr><th>الوقت</th><th>المشرف</th><th>الإجراء</th><th>العنصر</th><th>قبل</th><th>بعد</th></tr></thead><tbody>'+
    list.map(function(a){return '<tr><td class="muted">'+new Date(a.created_at).toLocaleString('ar-SA')+'</td><td>'+esc(a.actor||'')+'</td><td>'+esc(actionAr(a.action))+'</td><td>'+esc(entityAr(a.entity||''))+(a.entity_id?'#'+esc(a.entity_id):'')+'</td><td class="snip muted">'+esc((a.old_value||'').slice(0,40))+'</td><td class="snip">'+esc((a.new_value||'').slice(0,40))+'</td></tr>';}).join('')+
    '</tbody></table>';
}

// ---------- MODAL plumbing ----------
function openModal(){ document.getElementById('qg-modal').style.display='flex'; }
function closeModal(){ document.getElementById('qg-modal').style.display='none'; }
async function modalSave(){
  if(MODAL_MODE==='rule'||MODAL_MODE==='rule_chat'){
    var body={phrase:val('m_phrase'),scope:val('m_scope'),alert_type:val('m_type'),severity:val('m_sev'),ai_reason:val('m_reason'),suggested_correction:val('m_sugg'),_actor:ACTOR};
    if(!body.phrase||!body.alert_type){ alert('العبارة والنوع مطلوبان'); return; }
    var url = MODAL_ID? ('admin/rules/'+MODAL_ID) : 'admin/rules';
    await fetch(url,{method:MODAL_ID?'PUT':'POST',headers:H(),body:JSON.stringify(body)});
    closeModal(); loadRules();
  } else if(MODAL_MODE==='policy'){
    var pb={policy_category:val('m_cat'),official_statement:val('m_stmt'),numbers_or_limits:val('m_nums'),source_url:val('m_url'),_actor:ACTOR};
    if(MODAL_ID) pb.id=MODAL_ID;
    if(!pb.official_statement){ alert('النص الرسمي مطلوب'); return; }
    await fetch('admin/policies',{method:'POST',headers:H(),body:JSON.stringify(pb)});
    closeModal(); loadPolicies();
  }
}

// ---------- helpers ----------
function opts(vals,labels,cur){ return vals.map(function(v,i){return '<option value="'+v+'" '+(cur===v?'selected':'')+'>'+labels[i]+'</option>';}).join(''); }
function val(id){ var e=document.getElementById(id); return e? e.value : ''; }
function esc(x){ return (x||'').toString().replace(/</g,'&lt;'); }

</script>
</body></html>"""


# ---------- Section 1: official policy management ----------
from fastapi import Body

@router.get("/policies")
async def list_policies():
    p = await _pool()
    async with p.acquire() as c:
        rows = await c.fetch("SELECT * FROM qg_policies ORDER BY policy_category")
    return {"count": len(rows), "policies": [dict(r) for r in rows]}


@router.post("/policies")
async def upsert_policy(payload: dict = Body(...)):
    """Admin manual entry of an official policy (source of truth for section 1)."""
    import hashlib
    cat = (payload.get("policy_category") or "").strip()
    stmt = (payload.get("official_statement") or "").strip()
    if not cat or not stmt:
        return Response(content="policy_category and official_statement required", status_code=400)
    chash = hashlib.sha256(stmt.encode("utf-8")).hexdigest()[:16]
    p = await _pool()
    async with p.acquire() as c:
        existing = await c.fetchval("SELECT id FROM qg_policies WHERE policy_category=$1 AND is_active", cat)
        if existing:
            await c.execute("""UPDATE qg_policies SET official_statement=$2, numbers_or_limits=$3,
                   conditions=$4, exceptions=$5, source_url=$6, page_title=$7, content_hash=$8,
                   last_verified_at=now(), updated_at=now(), stale=FALSE WHERE id=$1""",
                existing, stmt, payload.get("numbers_or_limits"), payload.get("conditions"),
                payload.get("exceptions"), payload.get("source_url"), payload.get("page_title"), chash)
            return {"updated": existing, "category": cat}
        new_id = await c.fetchval("""INSERT INTO qg_policies
                (policy_category, official_statement, numbers_or_limits, conditions, exceptions,
                 source_url, page_title, content_hash, last_fetched_at)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,now()) RETURNING id""",
                cat, stmt, payload.get("numbers_or_limits"), payload.get("conditions"),
                payload.get("exceptions"), payload.get("source_url"), payload.get("page_title"), chash)
        return {"created": new_id, "category": cat}


@router.delete("/policies/{policy_id}")
async def deactivate_policy(policy_id: int):
    p = await _pool()
    async with p.acquire() as c:
        await c.execute("UPDATE qg_policies SET is_active=FALSE, updated_at=now() WHERE id=$1", policy_id)
    return {"deactivated": policy_id}


# ---------- Admin Settings (passphrase-gated, audited) ----------
import admin as _admin
from fastapi import Header

async def _require_admin(x_qg_admin: str = "", x_cw_token: str = "", x_cw_user: str = "",
                         x_cw_at: str = "", x_cw_client: str = "", x_cw_uid: str = ""):
    # primary: verify by the viewer's own Chatwoot session (devise-token-auth headers)
    if x_cw_at and x_cw_client and x_cw_uid:
        who = await _admin.verify_session_admin(x_cw_at, x_cw_client, x_cw_uid)
        if who:
            return who
    # secondary: Chatwoot admin via iframe currentAgent id
    if x_cw_user:
        who = await _admin.verify_admin_by_user_id(x_cw_user)
        if who:
            return who
    # alt: explicit Chatwoot token
    if x_cw_token:
        who = await _admin.verify_chatwoot_admin(x_cw_token)
        if who:
            return who
    # fallback: legacy passphrase (break-glass)
    if x_qg_admin and await _admin.verify_admin(x_qg_admin):
        return "passphrase-admin"
    return None

@router.post("/admin/login")
async def admin_login(payload: dict = Body(...)):
    # Preferred: verify by the viewer's Chatwoot session headers (from cw_d_session_info cookie)
    at = payload.get("access_token",""); cl = payload.get("client",""); ud = payload.get("uid","")
    if at and cl and ud:
        who = await _admin.verify_session_admin(at, cl, ud)
        return {"ok": bool(who), "actor": who, "mode": "session"}
    # iframe currentAgent id
    uid = payload.get("cw_user", "")
    if uid:
        who = await _admin.verify_admin_by_user_id(uid)
        return {"ok": bool(who), "actor": who, "mode": "chatwoot"}
    tok = payload.get("cw_token", "")
    if tok:
        who = await _admin.verify_chatwoot_admin(tok)
        return {"ok": bool(who), "actor": who, "mode": "chatwoot"}
    # fallback passphrase
    ok = await _admin.verify_admin(payload.get("passphrase",""))
    return {"ok": bool(ok), "actor": "passphrase-admin" if ok else None, "mode": "passphrase"}

@router.get("/admin/rules")
async def admin_rules(x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    return {"rules": await _admin.rules_list()}

@router.post("/admin/rules")
async def admin_rule_create(payload: dict = Body(...), x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    actor = payload.get("_actor","admin")
    nid = await _admin.rule_create(payload, actor)
    return {"created": nid}

@router.put("/admin/rules/{rid}")
async def admin_rule_update(rid: int, payload: dict = Body(...), x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    actor = payload.get("_actor","admin")
    ok = await _admin.rule_update(rid, payload, actor)
    return {"updated": ok}

@router.delete("/admin/rules/{rid}")
async def admin_rule_delete(rid: int, x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default=""), actor: str = Query(default="admin")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    await _admin.rule_delete(rid, actor)
    return {"deleted": rid}

@router.get("/admin/settings")
async def admin_settings(x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    p = await _pool()
    async with p.acquire() as c:
        rows = await c.fetch("SELECT key, value, updated_at FROM qg_settings WHERE key != 'admin_pass_hash' ORDER BY key")
    return {"settings": [dict(r) for r in rows]}

@router.put("/admin/settings/{key}")
async def admin_setting_update(key: str, payload: dict = Body(...), x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    await _admin.set_setting(key, payload.get("value",""), payload.get("_actor","admin"))
    return {"ok": True}

@router.get("/admin/audit")
async def admin_audit(x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    return {"audit": await _admin.audit_list()}


# ---------- Admin: policies (gated + audited) ----------
@router.get("/admin/policies")
async def admin_policies(x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    if not await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid):
        return Response(content="unauthorized", status_code=401)
    p = await _pool()
    async with p.acquire() as c:
        rows = await c.fetch("SELECT * FROM qg_policies ORDER BY policy_category")
    return {"policies": [dict(r) for r in rows]}

@router.post("/admin/policies")
async def admin_policy_save(payload: dict = Body(...), x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    actor = await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid)
    if not actor:
        return Response(content="unauthorized", status_code=401)
    import hashlib
    cat = (payload.get("policy_category") or "").strip()
    stmt = (payload.get("official_statement") or "").strip()
    if not cat or not stmt:
        return Response(content="missing fields", status_code=400)
    chash = hashlib.sha256(stmt.encode("utf-8")).hexdigest()[:16]
    pid = payload.get("id")
    p = await _pool()
    async with p.acquire() as c:
        if pid:
            old = await c.fetchval("SELECT official_statement FROM qg_policies WHERE id=$1", pid)
            await c.execute("""UPDATE qg_policies SET policy_category=$2, official_statement=$3, numbers_or_limits=$4,
                   conditions=$5, source_url=$6, content_hash=$7, updated_at=now(), stale=FALSE WHERE id=$1""",
                pid, cat, stmt, payload.get("numbers_or_limits"), payload.get("conditions"),
                payload.get("source_url"), chash)
            await _admin._audit(c, actor, "update_policy", "policies", pid, old, stmt[:120])
            return {"updated": pid}
        nid = await c.fetchval("""INSERT INTO qg_policies (policy_category, official_statement, numbers_or_limits, conditions, source_url, content_hash, last_fetched_at)
                VALUES ($1,$2,$3,$4,$5,$6,now()) RETURNING id""",
                cat, stmt, payload.get("numbers_or_limits"), payload.get("conditions"), payload.get("source_url"), chash)
        await _admin._audit(c, actor, "create_policy", "policies", nid, None, stmt[:120])
        return {"created": nid}

@router.put("/admin/policies/{pid}/toggle")
async def admin_policy_toggle(pid: int, x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default="")):
    actor = await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid)
    if not actor:
        return Response(content="unauthorized", status_code=401)
    p = await _pool()
    async with p.acquire() as c:
        cur = await c.fetchval("SELECT is_active FROM qg_policies WHERE id=$1", pid)
        newv = not cur
        await c.execute("UPDATE qg_policies SET is_active=$2, updated_at=now() WHERE id=$1", pid, newv)
        await _admin._audit(c, actor, "toggle_policy", "policies", pid, str(cur), str(newv))
    return {"id": pid, "is_active": newv}

@router.delete("/admin/policies/{pid}")
async def admin_policy_delete(pid: int, x_qg_admin: str = Header(default=""), x_cw_token: str = Header(default=""), x_cw_user: str = Header(default=""), x_cw_at: str = Header(default=""), x_cw_client: str = Header(default=""), x_cw_uid: str = Header(default=""), actor: str = Query(default="admin")):
    who = await _require_admin(x_qg_admin, x_cw_token, x_cw_user, x_cw_at, x_cw_client, x_cw_uid)
    if not who:
        return Response(content="unauthorized", status_code=401)
    p = await _pool()
    async with p.acquire() as c:
        await c.execute("UPDATE qg_policies SET is_active=FALSE, updated_at=now() WHERE id=$1", pid)
        await _admin._audit(c, who, "delete_policy", "policies", pid, None, None)
    return {"deleted": pid}
